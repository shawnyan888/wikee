#!/usr/bin/python
# -*- coding: utf-8 -*-
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook

__author__ = 'Shawn Yan'
__date__ = '10:11 2018/5/23'


class ExcelWriter(object):
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.wb = Workbook(excel_file)

    def get_sheet(self, sheet_name):
        ws = self.wb.get_worksheet_by_name(sheet_name)
        if not ws:
            ws = self.wb.add_worksheet(sheet_name)
        return ws

    def close(self):
        self.wb.close()


def merge_dict(new_dict, old_dict):
    new_dict.update(old_dict)
    return new_dict


class WriterHelper(ExcelWriter):
    """
    Base on http://xlsxwriter.readthedocs.io/format.html
    """
    def __init__(self, excel_file):
        super(WriterHelper, self).__init__(excel_file)

    def write_fonts(self):
        """Font type  'font_name'  set_font_name()
           Font size  'font_size'  set_font_size()
           Font color 'font_color' set_font_color()
           Bold       'bold'       set_bold()
           Italic     'italic'     set_italic()
           Underline  'underline'  set_underline()
           Strikeout    'font_strikeout' set_font_strikeout()
           Super/Subscript 'font_script' set_font_script()
        """
        my_font_property = dict(font_name="Arial",
                                font_color="red",
                                bold=True,
                                italic=True,
                                underline=True,
                                font_strikeout=2)
        ws = self.get_sheet("Fonts")
        font1 = self.wb.add_format(my_font_property)
        font2 = self.wb.add_format(merge_dict(dict(font_size=23), my_font_property))
        font3 = self.wb.add_format(dict(font_script=1))
        font4 = self.wb.add_format(merge_dict(dict(font_script=2), my_font_property))
        for row, _font in enumerate([font1, font2, font3, font4]):
            ws.write(row, 1, "Peter", _font)
            ws.write_rich_string(row, 3, "This is ", font1, "font1", font2, "font2",  font3, "font3",  font4, "font4")
        ws.set_column(1, 1, 80)

    def write_alignments(self):
        """Alignment Horizontal align 'align' set_align()  left,center,right,fill,justify,center_across,distributed
                     Vertical align   'valign' set_align() top,vcenter,bottom,vjustify,vdistributed
                     Rotation  'rotation' set_rotation()
                     Text wrap 'text_wrap' set_text_wrap()
                     Reading order 'reading_order' set_reading_order()
                     Justify last  'text_justlast' set_text_justlast()
                     Center across 'center_across' set_center_across()
                     Indentation   'indent' set_indent()
                     Shrink to fit 'shrink' set_shrink()
        """
        ws = self.get_sheet("Alignments")
        align1 = self.wb.add_format(dict(align="center", text_wrap=True))
        align2 = self.wb.add_format(dict(align="left"))
        _ = "Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday"
        ws.write("A2", _, align1)
        ws.write("A5", _, align2)

    def write_borders(self):
        """Cell border   'border'  set_border()
           Bottom border 'bottom'  set_bottom()
           Top border    'top'     set_top()
           Left border   'left'    set_left()
           Right border  'right'   set_right()
           Border color  'border_color'  set_border_color()
           Bottom color  'bottom_color'  set_bottom_color()
           Top color     'top_color'   set_top_color()
           Left color    'left_color'  set_left_color()
           Right color   'right_color' set_right_color()
        """
        ws = self.get_sheet("borders")
        formats = dict()
        for i in range(14):
            formats[i] = self.wb.add_format()
        for i in range(14):
            _format = formats.get(i)
            _format.set_border(i)
            ws.write(i*2, 2, "border_%d" % i, _format)
        ws.merge_range(10, 12, 10, 14, "check", formats[10])

    def write_filter1_first(self):
        ws = self.get_sheet("filter1")
        for i in range(50):
            for j in range(34):
                ii = "Peter%d" % i if i < 9 else i
                ws.write(i, j, "%s_%s" % (ii, j))
        ws.autofilter(3, 2, 3, 8)

    def write_filter2_first(self):
        ws = self.get_sheet("filter2")
        for i in range(50):
            for j in range(34):
                ii = "Peter%d" % i if i < 9 else i
                ws.write(i, j, "%s_%s" % (ii, j))
        ws.autofilter(0, 2, 0, 8)

    # Add Filter may change other cells format!!!
    # just a sample.
    def add_filter1(self):
        wb = load_workbook(self.excel_file)
        ws = wb["filter1"]
        ws.auto_filter.ref = "B:N"
        wb.save("filter1" + self.excel_file)

    def add_filter2(self):
        wb = load_workbook(self.excel_file)
        ws = wb["filter2"]
        ws.auto_filter.ref = "B8:N8"
        ws.cell(5, 5, "shawn")
        wb.save("filter2" + self.excel_file)

if __name__ == "__main__":
    hlp = WriterHelper("shawn.xlsx")
    hlp.write_fonts()
    hlp.write_alignments()
    hlp.write_borders()
    hlp.write_filter1_first()
    hlp.write_filter2_first()
    hlp.close()
    # hlp.add_filter1()
    # hlp.add_filter2()
